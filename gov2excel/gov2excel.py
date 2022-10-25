from time import strptime
from pynput import keyboard
from selenium.webdriver.support.ui import Select
import PublicDataReader as pdr
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import os
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime


os.system("cls")
print("Starting ...\n")
print("  ┏━")
print("   ※ 작업용 크롬이 켜지면 시작합니다.")
print("                                    ━┛")


options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()), options=options
)
os.system("cls")





# 엑셀 불러오기
try:
    f = open("./setting.txt",'r',encoding='utf-8')
    directory = f.read()
    if directory == '':
        raise
    print(f"경로 : {directory}")
    f.close()
except:
    f = open("./setting.txt",'w',encoding='utf-8')
    directory = input("엑셀 파일 경로를 입력해주세요(예시: C:\\Users\\User\\Documents\\부성지구 지번별 현황.xlsx) : ")
    f.write(directory)
    print(f"경로:{directory}")
    f.close()

wb = load_workbook(directory,data_only=True)
ws = wb['Sheet1']

# 빈칸 가져오고 주소와 몇번째인지 반환
arr = []
for i in range(3,476):
    if not ws.cell(i,5).value:
        arr.append([ws.cell(i,2).value,ws.cell(i,3).value,i])






def focus():
    pos = driver.get_window_position()
    driver.minimize_window()
    driver.set_window_position(pos['x'],pos['y'])

focus()

# 정부 사이트 접속
driver.get("https://www.gov.kr/nlogin/?Mcode=10003")
driver.find_element(By.CSS_SELECTOR, 'a[data-tg="match03"]').click()
main_window = driver.current_window_handle

# 로그인
def login(driver):
    idfield = driver.find_element(By.NAME, 'userId')
    pwdfield = driver.find_element(By.NAME, 'pwd')
    lgnbtn = driver.find_element(By.ID, 'genLogin')

    idfield.clear()
    idfield.send_keys("bhj0525")

    pwdfield.clear()
    pwdfield.send_keys("bang0525**")

    lgnbtn.click()
    driver.get("https://www.gov.kr")


login(driver)
print("로그인 성공")




for i in arr:
    driver.switch_to.window(main_window)
    if i[1] == '신당동':
        address = '천안시 신당동'
    else:
        address = i[0]
    bun = i[1]
    ji = '0'

    # 건축물대장 발급 신청 url 접속
    driver.get("https://www.gov.kr/mw/AA040OfferMainFrm.do?capp_biz_cd=13100000026&HighCtgCD=A02001001&FAX_TYPE=B&img=02&selectedSeq=01")
    main = driver.window_handles[0]
    focus()

    # 건축물대장 열람 신청 url 접속
    element = WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, "a[onclick=\"javascript:refreshFormRadio('03');\"]")))
    element.click()

    # 건축물 소재지 입력 팝업에서 검색

    def popupcontrol(driver):

        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="btnAddress"]')))
        # 핸들 가져오기
        handles = driver.window_handles

        element.click()
        try:
            WebDriverWait(driver,10).until(EC.new_window_is_opened(handles))
        except:
            pass

        # 핸들 가져오기
        handles2 = driver.window_handles

        # 팝업 핸들 구하기
        popup_window = handles2[-1]

        # 팝업 핸들로 스위치
        driver.switch_to.window(popup_window)
        focus()

        p_searchfield = driver.find_element(By.NAME, 'txtAddr')
        p_searchfield.clear()
        p_searchfield.send_keys(address)

        p_searchbtn = driver.find_element(
            By.CSS_SELECTOR, "button[onclick=\"isValid();return false;\"]")
        p_searchbtn.click()

        return popup_window

    popup_window = popupcontrol(driver)

    # 팝업에서 검색 결과 몇개인지 확인
    def len_p_elements(driver):
        try:
            p_wait = WebDriverWait(driver, 3).until(EC.element_to_be_clickable(
                (By.XPATH, '(//div[@id="resultList"]/a[@title="주소"])[1]')))
            p_elements = driver.find_elements(
                By.XPATH, '//div[@id="resultList"]/a[@title="주소"]')
        except:
            return []

        return p_elements

    search = len_p_elements(driver)
    p_len = len(search)


    # 검색결과 처리
    if (p_len == 0):
        print("결과 없음")
        driver.switch_to.window(popup_window)
        driver.close()
        continue

    elif (p_len > 1):
        print("결과가 2개 이상입니다.")
        results = driver.find_elements(By.XPATH,'//div[@id="resultList"]/a[@title="주소"]/dl/dd/div')
        pos = driver.get_window_position()
        driver.minimize_window()

        while(1):
            for i in range(p_len):
                print(f"{i+1} : {results[i].text}")
            try:
                search_input = int(input('입력 : '))-1
            except:
                continue
            
            if(search_input not in range(p_len)):
                print('잘못된 입력입니다.\n')
                continue
            else:
                print(f"선택 : '{results[search_input].text}'")
                results[search_input].click()
                break
    
    elif (p_len == 1):
        driver.find_element(By.XPATH, '//div[@id="resultList"]/a[@title="주소"]/dl/dd/div').click()


    driver.switch_to.window(main_window)
    focus()

    # main window에 작성
    driver.find_element(By.XPATH,'//input[@title="번지수"]').send_keys(bun)
    driver.find_element(By.XPATH,'//input[@title="호수"]').send_keys(ji)

    # submit
    driver.execute_script("setSinchungGubun('1');__call()")

    # 대기
    try:
        element = WebDriverWait(driver,40).until(EC.element_to_be_clickable((By.NAME,'issue_img')))
    except:
        print("오류")
        continue

    # 핸들 가져오기
    handles = driver.window_handles

    element.click()
    
    WebDriverWait(driver,10).until(EC.new_window_is_opened(handles))

    # 핸들 가져오기
    handles2 = driver.window_handles

    # 팝업 핸들 구하기
    popup_window2 = handles2[-1]

    # 팝업 핸들로 스위치
    driver.switch_to.window(popup_window2)
    focus()

    # 정보 가져오기
    documents_area = driver.find_element(By.XPATH,'(//table[@class="L2"])[1]/tbody/tr[4]/td[2]/span').text
    documents_owner = driver.find_element(By.XPATH,'(//table[@class="L2"])[1]/tbody/tr[5]/td[2]').text
    documents_address = driver.find_element(By.XPATH,'(//table[@class="L2"])[1]/tbody/tr[4]/td[5]').text
    documents_number = driver.find_element(By.XPATH,'(//table[@class="L2"])[1]/tbody/tr[5]/td[3]').text
    # 별 지우기
    documents_number = re.sub(r'\*+','',documents_number)

    documents_year = driver.find_element(By.XPATH,'(//table[@class="L2"])[1]/tbody/tr[4]/td[4]').text
    # 날짜 형식 변환
    documents_year = datetime.strptime(documents_year,'%Y년 %m월 %d일')
    documents_year = datetime.strftime(documents_year,'%Y/%m/%d')

    documents_reason = driver.find_element(By.XPATH,'(//table[@class="L2"])[1]/tbody/tr[5]/td').text
    # 괄호 지우기
    documents_reason = re.sub(r'\([^)]*\)', '', documents_reason)

    # 팝업 닫기
    driver.close()

    # 엑셀 쓰기
    ws.cell(i[2],5).value = float(documents_area)
    ws.cell(i[2],7).value = documents_owner
    ws.cell(i[2],8).value = documents_address
    ws.cell(i[2],9).value = documents_number
    ws.cell(i[2],11).value = documents_year
    ws.cell(i[2],12).value = documents_reason
    
    # 엑셀 정렬
    ws.cell(i[2],5).alignment = Alignment(horizontal='center')
    ws.cell(i[2],7).alignment = Alignment(horizontal='left')
    ws.cell(i[2],8).alignment = Alignment(horizontal='left')
    ws.cell(i[2],9).alignment = Alignment(horizontal='left')
    ws.cell(i[2],11).alignment = Alignment(horizontal='center')
    ws.cell(i[2],12).alignment = Alignment(horizontal='center')

    wb.save(directory)









